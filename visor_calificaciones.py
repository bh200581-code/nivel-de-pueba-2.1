"""
visor_calificaciones.py — Auditor de Calificaciones ETP (CORREGIDO + MEJORADO)
• Lee asignaciones reales desde core.auth (modulos_docentes).
• FIX: KeyError notas_json, formato sobre vacíos, sin-calificar vs reprobado, float seguro.
• NUEVO: Semáforo por R.A. (tasa de aprobación) + KPI "Sin calificar".
• Lógica intacta: cascada Ev/R1/R2/R3, matriz MINERD en Excel y recuperación.
"""
import json
import sqlite3
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core import auth

# ═══════════════════════════════════════════════════════════════════════════
# NÚCLEO DE DATOS
# ═══════════════════════════════════════════════════════════════════════════
DB_NAME = "gestion_etp.db"
INTENTOS = ["Ev", "R1", "R2", "R3"]


def asegurar_tablas():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS config_ra (
        modulo TEXT, seccion TEXT, cantidad_ra INTEGER, pesos_json TEXT,
        PRIMARY KEY (modulo, seccion)
    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS calificaciones (
        modulo TEXT, seccion TEXT, estudiante TEXT, notas_json TEXT,
        PRIMARY KEY (modulo, seccion, estudiante)
    )''')
    conn.commit()
    conn.close()


def obtener_config_ra(modulo: str, seccion: str) -> tuple:
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT cantidad_ra, pesos_json FROM config_ra WHERE modulo=? AND seccion=?", (modulo, seccion))
    row = cursor.fetchone()
    conn.close()
    if row:
        return row[0], json.loads(row[1])
    return 0, {}


def obtener_notas_df(modulo: str, seccion: str) -> pd.DataFrame:
    conn = sqlite3.connect(DB_NAME)
    try:
        df = pd.read_sql_query(
            "SELECT estudiante, notas_json FROM calificaciones WHERE modulo=? AND seccion=?",
            conn, params=(modulo, seccion),
        )
    except Exception:
        df = pd.DataFrame(columns=["estudiante", "notas_json"])
    conn.close()
    return df


def construir_df_notas(df_raw: pd.DataFrame) -> pd.DataFrame:
    if df_raw.empty:
        return pd.DataFrame()
    datos = []
    for _, r in df_raw.iterrows():
        fila = {"Estudiante": r['estudiante']}
        if r['notas_json']:
            fila.update(json.loads(r['notas_json']))
        datos.append(fila)
    return pd.DataFrame(datos)


def to_float(v):
    if v is None:
        return None
    try:
        if str(v).strip() == "" or pd.isna(v):
            return None
        return float(v)
    except Exception:
        return None


def tiene_notas(row, pesos_ra: dict) -> bool:
    """True si el estudiante tiene al menos una nota válida en cualquier RA/intento."""
    for ra in pesos_ra.keys():
        for intento in INTENTOS:
            v = to_float(row.get(f"{ra} - {intento}"))
            if v is not None:
                return True
    return False


def calcular_nota_final_fila(row, pesos_ra: dict) -> float:
    total = 0.0
    for ra in pesos_ra.keys():
        valores = [to_float(row.get(f"{ra} - {intento}")) for intento in INTENTOS]
        valores_validos = [v for v in valores if v is not None]
        if valores_validos:
            total += valores_validos[-1]
    return round(total, 1)


def generar_reporte_estados(df_raw: pd.DataFrame, pesos_ra: dict) -> pd.DataFrame:
    if df_raw.empty:
        return pd.DataFrame()
    reporte = []
    for _, row in df_raw.iterrows():
        estudiante = row['estudiante']
        notas = json.loads(row['notas_json']) if row['notas_json'] else {}   # ← FIX KeyError
        for ra, peso_total in pesos_ra.items():
            umbral_70 = peso_total * 0.7
            ev = to_float(notas.get(f"{ra} - Ev"))
            r1 = to_float(notas.get(f"{ra} - R1"))
            r2 = to_float(notas.get(f"{ra} - R2"))
            r3 = to_float(notas.get(f"{ra} - R3"))
            estado, etapa, nota_vigente = "Sin Evaluar", "-", 0
            if ev is None and r1 is None and r2 is None and r3 is None:
                pass
            else:
                if r3 is not None:
                    estado, etapa, nota_vigente = ("Recuperado", "R3", r3) if r3 >= umbral_70 else ("Reprobado", "R3 (Agotado)", r3)
                elif r2 is not None:
                    estado, etapa, nota_vigente = ("Recuperado", "R2", r2) if r2 >= umbral_70 else ("Pendiente", "R2 (Fallido)", r2)
                elif r1 is not None:
                    estado, etapa, nota_vigente = ("Recuperado", "R1", r1) if r1 >= umbral_70 else ("Pendiente", "R1 (Fallido)", r1)
                elif ev is not None:
                    estado, etapa, nota_vigente = ("Aprobado", "Ev", ev) if ev >= umbral_70 else ("Pendiente", "Ev (Fallido)", ev)
            reporte.append({
                "Estudiante": estudiante, "R.A.": ra, "Peso RA": peso_total,
                "Umbral (70%)": round(umbral_70, 1),
                "Nota Vigente": nota_vigente if nota_vigente > 0 else "-",
                "Estatus": estado, "Etapa": etapa,
            })
    return pd.DataFrame(reporte)


def calcular_semaforo_ra(df_raw: pd.DataFrame, pesos_ra: dict) -> pd.DataFrame:
    """NUEVO: tasa de aprobación por R.A. para el semáforo."""
    rows = []
    for ra, peso in pesos_ra.items():
        umbral = peso * 0.7
        aprob = pend = sin = 0
        for _, row in df_raw.iterrows():
            notas = json.loads(row['notas_json']) if row['notas_json'] else {}
            val = None
            for intento in INTENTOS:
                v = to_float(notas.get(f"{ra} - {intento}"))
                if v is not None:
                    val = v
            if val is None:
                sin += 1
            elif val >= umbral:
                aprob += 1
            else:
                pend += 1
        rows.append({"R.A.": ra, "Aprobados": aprob, "Pendientes": pend, "Sin nota": sin})
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL OFICIAL MINERD (float seguro)
# ═══════════════════════════════════════════════════════════════════════════
def generar_excel_matriz(df_notas, pesos_ra, seccion_nombre):
    wb = Workbook()
    ws = wb.active
    ws.title = str(seccion_nombre)[:31]
    fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_green = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fill_red = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')
    align_vert = Alignment(horizontal='center', vertical='center', textRotation=90)
    font_bold = Font(bold=True)
    font_bold_large = Font(bold=True, size=16)

    ws.cell(row=1, column=1, value=" ").fill = fill_gray
    ws.cell(row=1, column=2, value=" ").fill = fill_yellow
    ws.cell(row=2, column=1, value=" ").fill = fill_gray
    ws.cell(row=2, column=2, value=" ").fill = fill_yellow
    ws.cell(row=3, column=1, value="#").font = font_bold
    ws.cell(row=3, column=1).fill = fill_gray
    ws.cell(row=3, column=2, value="Nombres").font = font_bold
    ws.cell(row=3, column=2).fill = fill_green

    col_idx = 3
    suma_pesos = 0
    ra_col_ranges = []
    for ra_name, peso in pesos_ra.items():
        ws.cell(row=1, column=col_idx, value=f"{ra_name} ({peso})").font = font_bold
        ws.cell(row=1, column=col_idx).alignment = align_center
        ws.cell(row=1, column=col_idx).fill = fill_gray
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 3)
        ws.cell(row=2, column=col_idx, value=peso).font = font_bold_large
        ws.cell(row=2, column=col_idx).alignment = align_center
        ws.cell(row=2, column=col_idx).fill = fill_gray
        ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 3)
        for i, intento in enumerate(INTENTOS):
            c = ws.cell(row=3, column=col_idx + i, value=intento)
            c.font = Font(bold=True, size=9)
            c.alignment = align_center
            c.fill = fill_gray
        for r in range(1, 4):
            for c in range(col_idx, col_idx + 4):
                ws.cell(row=r, column=c).border = border_thin
        ra_col_ranges.append((col_idx, col_idx + 3, peso))
        col_idx += 4
        suma_pesos += peso

    col_total = col_idx
    ws.cell(row=1, column=col_total, value="Total").font = font_bold
    ws.cell(row=1, column=col_total).alignment = align_center
    ws.cell(row=1, column=col_total).fill = fill_gray
    ws.cell(row=2, column=col_total, value=suma_pesos).font = font_bold_large
    ws.cell(row=2, column=col_total).alignment = align_center
    ws.cell(row=2, column=col_total).fill = fill_gray
    ws.merge_cells(start_row=2, start_column=col_total, end_row=3, end_column=col_total)

    col_sit = col_idx + 1
    ws.cell(row=1, column=col_sit, value="Situacion Final").font = Font(bold=True, size=9)
    ws.cell(row=1, column=col_sit).alignment = align_center
    ws.cell(row=1, column=col_sit).fill = fill_gray
    ws.merge_cells(start_row=1, start_column=col_sit, end_row=1, end_column=col_sit + 1)
    ws.cell(row=2, column=col_sit, value="Aprobado").font = font_bold
    ws.cell(row=2, column=col_sit).alignment = align_vert
    ws.cell(row=2, column=col_sit).fill = fill_green
    ws.merge_cells(start_row=2, start_column=col_sit, end_row=3, end_column=col_sit)
    ws.cell(row=2, column=col_sit + 1, value="Reprobado").font = font_bold
    ws.cell(row=2, column=col_sit + 1).alignment = align_vert
    ws.cell(row=2, column=col_sit + 1).fill = fill_red
    ws.merge_cells(start_row=2, start_column=col_sit + 1, end_row=3, end_column=col_sit + 1)
    for r in range(1, 4):
        for c in range(1, col_sit + 2):
            if not ws.cell(row=r, column=c).border:
                ws.cell(row=r, column=c).border = border_thin

    row_idx = 4
    num_students = len(df_notas) if not df_notas.empty else 15
    for i in range(1, num_students + 1):
        ws.cell(row=row_idx, column=1, value=i).border = border_thin
        ws.cell(row=row_idx, column=1).alignment = align_center
        estudiante = " "
        if not df_notas.empty and i - 1 < len(df_notas):
            estudiante = df_notas.iloc[i - 1].get('Estudiante', '')
        ws.cell(row=row_idx, column=2, value=estudiante).border = border_thin
        c_idx = 3
        if not df_notas.empty and i - 1 < len(df_notas):
            row_data = df_notas.iloc[i - 1]
            for ra in pesos_ra.keys():
                for intento in INTENTOS:
                    val = row_data.get(f"{ra} - {intento}", " ")
                    fv = to_float(val)          # ← FIX float seguro
                    val = " " if fv is None else int(round(fv))
                    c = ws.cell(row=row_idx, column=c_idx, value=val)
                    c.border = border_thin
                    c.alignment = align_center
                    c.number_format = '0'
                    c_idx += 1
        else:
            for _ in range(len(pesos_ra) * 4):
                c = ws.cell(row=row_idx, column=c_idx, value=" ")
                c.border = border_thin
                c.alignment = align_center
                c_idx += 1
        formula_parts = []
        for start_c, end_c, _ in ra_col_ranges:
            c_ev = get_column_letter(start_c) + str(row_idx)
            c_r1 = get_column_letter(start_c + 1) + str(row_idx)
            c_r2 = get_column_letter(start_c + 2) + str(row_idx)
            c_r3 = get_column_letter(start_c + 3) + str(row_idx)
            part = f"IF(ISNUMBER({c_r3}),{c_r3},IF(ISNUMBER({c_r2}),{c_r2},IF(ISNUMBER({c_r1}),{c_r1},IF(ISNUMBER({c_ev}),{c_ev},0))))"
            formula_parts.append(part)
        total_formula = "=SUM(" + ", ".join(formula_parts) + ")" if formula_parts else "=0"
        c_tot = ws.cell(row=row_idx, column=col_total, value=total_formula)
        c_tot.border = border_thin
        c_tot.alignment = align_center
        c_tot.font = font_bold
        c_tot.number_format = '0'
        tot_cell_ref = get_column_letter(col_total) + str(row_idx)
        c_apr = ws.cell(row=row_idx, column=col_sit, value=f'=IF({tot_cell_ref}>=70,"X"," ")')
        c_apr.border = border_thin
        c_apr.alignment = align_center
        c_rep = ws.cell(row=row_idx, column=col_sit + 1, value=f'=IF(AND({tot_cell_ref}<70,{tot_cell_ref}>0),"X"," ")')
        c_rep.border = border_thin
        c_rep.alignment = align_center
        row_idx += 1

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 35
    for c in range(3, col_total):
        ws.column_dimensions[get_column_letter(c)].width = 5.5
    ws.column_dimensions[get_column_letter(col_total)].width = 8
    ws.column_dimensions[get_column_letter(col_sit)].width = 6
    ws.column_dimensions[get_column_letter(col_sit + 1)].width = 6
    ws.row_dimensions[2].height = 45
    ws.row_dimensions[3].height = 20

    fill_rojo_alerta = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    font_roja_alerta = Font(color='9C0006', bold=True)
    for start_c, end_c, peso in ra_col_ranges:
        umbral = peso * 0.7
        rule = CellIsRule(operator='lessThan', formula=[str(umbral)], stopIfTrue=True, fill=fill_rojo_alerta, font=font_roja_alerta)
        rango_str = f"{get_column_letter(start_c)}4:{get_column_letter(end_c)}{row_idx - 1}"
        ws.conditional_formatting.add(rango_str, rule)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ═══════════════════════════════════════════════════════════════════════════
# VISTA EN PANTALLA (coerción numérica segura)
# ═══════════════════════════════════════════════════════════════════════════
def _col_o_nan(df, col):
    if col in df.columns:
        return pd.to_numeric(df[col], errors="coerce")
    return pd.Series(np.nan, index=df.index)


def construir_vista_matriz(df: pd.DataFrame, pesos_ra: dict):
    tuples = [(" ", "Nombres")]
    datos = {(" ", "Nombres"): df["Estudiante"] if "Estudiante" in df.columns else pd.Series(dtype=object)}
    for ra, peso in pesos_ra.items():
        for intento in INTENTOS:
            col = f"{ra} - {intento}"
            clave = (f"{ra} ({peso})", intento)
            tuples.append(clave)
            datos[clave] = _col_o_nan(df, col)
    clave_final = (" ", "Nota Final")
    tuples.append(clave_final)
    datos[clave_final] = _col_o_nan(df, "Nota Final")
    df_vista = pd.DataFrame(datos)
    df_vista.columns = pd.MultiIndex.from_tuples(tuples)

    def resaltar(data):
        estilos = pd.DataFrame("", index=data.index, columns=data.columns)
        for ra, peso in pesos_ra.items():
            umbral = peso * 0.7
            for intento in INTENTOS:
                col = (f"{ra} ({peso})", intento)
                if col in data.columns:
                    val = pd.to_numeric(data[col], errors="coerce")
                    mask = val.notna() & (val < umbral)
                    estilos.loc[mask, col] = "background-color: #FEE2E2; color: #DC2626; font-weight: bold;"
        if clave_final in data.columns:
            val_final = pd.to_numeric(data[clave_final], errors="coerce")
            mask_final = val_final.notna() & (val_final < 70)
            estilos.loc[mask_final, clave_final] = "background-color: #FECACA; color: #991B1B; font-weight: bold;"
        return estilos

    formato = {c: "{:.0f}" for c in df_vista.columns if c[1] in ("Ev", "R1", "R2", "R3", "Nota Final")}
    return df_vista.style.apply(resaltar, axis=None).format(formato, na_rep="")


# ═══════════════════════════════════════════════════════════════════════════
# ESTILOS
# ═══════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; background-color: #F8FAFC; color: #1E293B; }
.visor-hero { background: linear-gradient(135deg, #0F172A 0%, #1E3A8A 60%, #7C3AED 100%); color: #fff;
padding: 1.8rem; border-radius: 18px; margin-bottom: 1.25rem; box-shadow: 0 18px 35px rgba(15,23,42,.18); }
.visor-title { font-size: 2rem; font-weight: 800; letter-spacing: -0.02em; }
.visor-sub { opacity: .88; font-size: 1rem; margin-top: .3rem; }
.section-title { color: #1D4ED8; font-weight: 700; font-size: 1.12rem; border-bottom: 2px solid #DBEAFE;
padding-bottom: 8px; margin: 1.2rem 0 .9rem 0; }
.kpi-card { background:#fff; border:1px solid #E2E8F0; border-top:4px solid #2563EB; border-radius:12px;
padding:14px 16px; box-shadow:0 4px 12px rgba(15,23,42,.06); text-align:center; }
.kpi-value { font-size:2rem; font-weight:800; color:#0F172A; }
.kpi-label { font-size:.78rem; font-weight:800; color:#64748B; text-transform:uppercase; letter-spacing:.05em; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════
# GUARDIA + CACHÉ
# ═══════════════════════════════════════════════════════════════════════════
if not st.session_state.get("coordinador_autenticado", False):
    st.error("🔒 Esta página es exclusiva de Coordinación.")
    st.stop()

asegurar_tablas()


@st.cache_data(ttl=60, show_spinner=False)
def cargar_modulos_cached() -> pd.DataFrame:
    regs = auth.listar_asignaciones()
    if not regs:
        return pd.DataFrame(columns=["docente", "modulo", "seccion"])
    return pd.DataFrame(regs)[["docente", "modulo", "seccion"]]


@st.cache_data(ttl=30, show_spinner=False)
def cargar_config_cached(modulo: str, seccion: str):
    return obtener_config_ra(modulo, seccion)


@st.cache_data(ttl=30, show_spinner=False)
def cargar_notas_raw_cached(modulo: str, seccion: str) -> pd.DataFrame:
    return obtener_notas_df(modulo, seccion)


# ═══════════════════════════════════════════════════════════════════════════
# HERO + BARRA
# ═══════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="visor-hero">
    <div class="visor-title">👁️ Auditor de Calificaciones ETP</div>
    <div class="visor-sub">Monitoreo institucional en vivo · cálculo en cascada · recuperación · descarga oficial MINERD</div>
</div>
""", unsafe_allow_html=True)

col_title, col_refresh = st.columns([4, 1])
with col_title:
    n_doc = cargar_modulos_cached()["docente"].nunique()
    n_mod = len(cargar_modulos_cached())
    st.caption(f"🔄 Sincronizado con Gestor de Accesos y Directorio · {n_doc} docentes / {n_mod} asignaciones")
with col_refresh:
    if st.button("🔄 Actualizar datos", width="stretch", type="primary"):
        st.cache_data.clear()
        st.toast("Datos actualizados correctamente", icon="✅")
        st.rerun()

# ═══════════════════════════════════════════════════════════════════════════
# FILTROS: DOCENTE → MÓDULO (comportamiento original restaurado)
# ═══════════════════════════════════════════════════════════════════════════
df_modulos = cargar_modulos_cached()
if df_modulos.empty:
    st.warning("⚠️ No hay módulos registrados. Asigna módulos desde el Directorio o el Gestor de Accesos.")
    st.stop()

col1, col2 = st.columns(2)
with col1:
    docente_sel = st.selectbox(
        "1. Filtrar por docente",
        ["Todos los docentes"] + sorted(df_modulos["docente"].unique().tolist()),
        key="sel_docente",
    )

# 🔑 CLAVE: filtrar los módulos por el docente elegido ANTES de llenar el 2.º selector
df_filtrado = df_modulos.copy()
if docente_sel != "Todos los docentes":
    df_filtrado = df_filtrado[df_filtrado["docente"] == docente_sel]

with col2:
    modulos_list = sorted(
        (df_filtrado["modulo"] + " | " + df_filtrado["seccion"]).unique().tolist()
    )
    # Si el módulo seleccionado ya no pertenece al docente elegido, resetear
    if modulos_list and st.session_state.get("sel_modulo") not in modulos_list:
        st.session_state["sel_modulo"] = modulos_list[0]
    if not modulos_list:
        st.info("No hay módulos para el filtro seleccionado.")
        st.stop()
    mod_sec = st.selectbox(
        "2. Seleccionar módulo y sección",
        modulos_list,
        key="sel_modulo",
    )

modulo, seccion = mod_sec.split(" | ")
cant_ra, pesos_ra = cargar_config_cached(modulo, seccion)
if cant_ra == 0 or not pesos_ra:
    st.warning("⚠️ El docente aún no ha configurado sus R.A. para este módulo.")
    st.stop()

df_notas_raw = cargar_notas_raw_cached(modulo, seccion)
df_final = construir_df_notas(df_notas_raw)
if not df_final.empty:
    df_final["Nota Final"] = df_final.apply(lambda row: calcular_nota_final_fila(row, pesos_ra), axis=1)
else:
    df_final = pd.DataFrame(columns=["Estudiante", "Nota Final"])

df_reporte = generar_reporte_estados(df_notas_raw, pesos_ra)

# ── KPIs corregidos (sin-calificar separado) ──
total_estudiantes = 0 if df_final.empty else len(df_final)
if df_final.empty:
    aprobados = reprobados = sin_calificar = 0
else:
    mask_notas = df_final.apply(lambda r: tiene_notas(r, pesos_ra), axis=1)
    sin_calificar = int((~mask_notas).sum())
    df_con_notas = df_final[mask_notas]
    aprobados = int((df_con_notas["Nota Final"] >= 70).sum())
    reprobados = int((df_con_notas["Nota Final"] < 70).sum())
alertas_activas = 0 if df_reporte.empty else int(df_reporte["Estatus"].str.contains("Pendiente|Reprobado", case=False, na=False).sum())

k1, k2, k3, k4, k5 = st.columns(5)
with k1:
    st.markdown(f'<div class="kpi-card"><div class="kpi-label">👥 Estudiantes</div><div class="kpi-value">{total_estudiantes}</div></div>', unsafe_allow_html=True)
with k2:
    st.markdown(f'<div class="kpi-card" style="border-top-color:#10B981;"><div class="kpi-label">✅ Aprobados</div><div class="kpi-value" style="color:#059669;">{aprobados}</div></div>', unsafe_allow_html=True)
with k3:
    st.markdown(f'<div class="kpi-card" style="border-top-color:#EF4444;"><div class="kpi-label">❌ Reprobados</div><div class="kpi-value" style="color:#DC2626;">{reprobados}</div></div>', unsafe_allow_html=True)
with k4:
    st.markdown(f'<div class="kpi-card" style="border-top-color:#94A3B8;"><div class="kpi-label">⬜ Sin calificar</div><div class="kpi-value" style="color:#64748B;">{sin_calificar}</div></div>', unsafe_allow_html=True)
with k5:
    st.markdown(f'<div class="kpi-card" style="border-top-color:#F59E0B;"><div class="kpi-label">🚨 Alertas</div><div class="kpi-value" style="color:#D97706;">{alertas_activas}</div></div>', unsafe_allow_html=True)

tab_matriz, tab_recuperacion, tab_export = st.tabs(["📊 Matriz oficial", "🚨 Recuperación", "📥 Exportar"])

# ═══════════════════════════════════════════════════════════════════════════
# TAB 1: MATRIZ
# ═══════════════════════════════════════════════════════════════════════════
with tab_matriz:
    st.markdown('<div class="section-title">📊 Sábana de notas oficial</div>', unsafe_allow_html=True)
    if df_final.empty:
        st.info("📝 Aún no hay calificaciones registradas para este módulo.")
    else:
        st.caption("Las casillas en rojo no alcanzan el 70% del valor del R.A.")
        st.dataframe(construir_vista_matriz(df_final, pesos_ra), width="stretch", hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════
# TAB 2: RECUPERACIÓN + SEMÁFORO POR R.A.
# ═══════════════════════════════════════════════════════════════════════════
with tab_recuperacion:
    st.markdown('<div class="section-title">🚨 Seguimiento de recuperación por R.A.</div>', unsafe_allow_html=True)
    if df_reporte.empty:
        st.info("No hay datos suficientes para generar el estado de recuperación.")
    else:
        # NUEVO: Semáforo por R.A.
        df_semaforo = calcular_semaforo_ra(df_notas_raw, pesos_ra)
        st.markdown("#### 🚦 Semáforo por Resultado de Aprendizaje")
        st.bar_chart(df_semaforo.set_index("R.A.")[["Aprobados", "Pendientes"]])

        col_f1, col_f2 = st.columns([1, 2])
        with col_f1:
            filtro_estatus = st.selectbox("Filtrar por estatus", ["Todos", "Recuperado", "Pendiente", "Aprobado", "Reprobado", "Sin Evaluar"])
        with col_f2:
            busqueda = st.text_input("🔎 Buscar estudiante", placeholder="Nombre o parte del nombre")

        df_rep = df_reporte.copy()
        if filtro_estatus != "Todos":
            df_rep = df_rep[df_rep["Estatus"].str.contains(filtro_estatus, case=False, na=False)]
        if busqueda.strip():
            df_rep = df_rep[df_rep["Estudiante"].astype(str).str.contains(busqueda.strip(), case=False, na=False)]

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("Registros", len(df_rep))
        with c2:
            st.metric("Recuperados", int(df_rep["Estatus"].str.contains("Recuperado", na=False).sum()))
        with c3:
            st.metric("Pendientes", int(df_rep["Estatus"].str.contains("Pendiente", na=False).sum()))
        with c4:
            st.metric("Reprobados", int(df_rep["Estatus"].str.contains("Reprobado", na=False).sum()))

        st.markdown("#### Distribución por estatus")
        st.bar_chart(df_rep["Estatus"].value_counts())
        st.dataframe(df_rep, width="stretch", hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════
# TAB 3: EXPORTAR
# ═══════════════════════════════════════════════════════════════════════════
with tab_export:
    st.markdown('<div class="section-title">📥 Exportación institucional</div>', unsafe_allow_html=True)
    col_e1, col_e2 = st.columns(2)
    with col_e1:
        st.markdown("#### 📊 Matriz oficial MINERD")
        st.caption("Descarga la matriz completa con formato oficial, fórmulas y formato condicional.")
        if st.button("🧾 Generar Excel de matriz", width="stretch"):
            buffer_matriz = generar_excel_matriz(df_final, pesos_ra, seccion)
            st.toast("Matriz oficial generada", icon="🧾")
            st.download_button(
                label="📥 Descargar matriz (.xlsx)",
                data=buffer_matriz,
                file_name=f"Matriz_Oficial_{modulo}_{seccion}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                width="stretch",
            )
    with col_e2:
        st.markdown("#### 🚨 Reporte de recuperación")
        st.caption("Estado actual por estudiante y R.A., en Excel o CSV.")
        if df_reporte.empty:
            st.info("No hay reporte disponible.")
        else:
            buffer_rep = BytesIO()
            df_reporte.to_excel(buffer_rep, index=False, engine="openpyxl")
            buffer_rep.seek(0)
            st.download_button(
                label="📥 Descargar reporte (.xlsx)",
                data=buffer_rep,
                file_name=f"Estatus_Recuperacion_{modulo}_{seccion}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
            csv = df_reporte.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                label="📄 Descargar reporte (.csv)",
                data=csv,
                file_name=f"Estatus_Recuperacion_{modulo}_{seccion}.csv",
                mime="text/csv",
                width="stretch",
            )