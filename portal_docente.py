"""
portal_docente.py
=================
Portal del Docente ETP mejorado.
100% INDEPENDIENTE (Sin dependencias externas).
"""

import sqlite3
import pandas as pd
import json
import hashlib
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import streamlit as st

# ===========================================================================
# FUNCIONES NÚCLEO (Bases de datos, Lógica y Excel)
# ===========================================================================
DB_NAME = "gestion_etp.db"
INTENTOS = ['Ev', 'R1', 'R2', 'R3']

def asegurar_tablas():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS docentes (
                        docente TEXT, modulo TEXT, seccion TEXT, password TEXT DEFAULT '1234', usuario TEXT
                      )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS config_ra (
                        modulo TEXT, seccion TEXT, cantidad_ra INTEGER, pesos_json TEXT,
                        PRIMARY KEY (modulo, seccion)
                    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS calificaciones (
                        modulo TEXT, seccion TEXT, estudiante TEXT, notas_json TEXT,
                        PRIMARY KEY (modulo, seccion, estudiante)
                    )''')
    conn.commit()
    conn.close()

def obtener_config_ra(modulo: str, seccion: str) -> tuple:
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT cantidad_ra, pesos_json FROM config_ra WHERE modulo=? AND seccion=?", (modulo, seccion))
    row = cursor.fetchone()
    conn.close()
    if row: return row[0], json.loads(row[1])
    return 0, {}

def guardar_config_ra(modulo: str, seccion: str, cantidad: int, pesos_dict: dict) -> None:
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('INSERT OR REPLACE INTO config_ra (modulo, seccion, cantidad_ra, pesos_json) VALUES (?, ?, ?, ?)', 
                   (modulo, seccion, cantidad, json.dumps(pesos_dict)))
    conn.commit()
    conn.close()

def obtener_notas_df(modulo: str, seccion: str) -> pd.DataFrame:
    conn = sqlite3.connect(DB_NAME)
    try:
        df = pd.read_sql_query("SELECT estudiante, notas_json FROM calificaciones WHERE modulo=? AND seccion=?", conn, params=(modulo, seccion))
    except Exception:
        df = pd.DataFrame(columns=["estudiante", "notas_json"])
    conn.close()
    return df

def construir_df_notas(df_raw: pd.DataFrame) -> pd.DataFrame:
    if df_raw.empty: return pd.DataFrame()
    datos = []
    for _, r in df_raw.iterrows():
        fila = {"Estudiante": r['estudiante']}
        if r['notas_json']:
            fila.update(json.loads(r['notas_json']))
        datos.append(fila)
    return pd.DataFrame(datos)

def calcular_nota_final_fila(row, pesos_ra: dict) -> float:
    total = 0.0
    for ra in pesos_ra.keys():
        valores = [row.get(f"{ra} - {intento}") for intento in INTENTOS]
        valores_validos = [v for v in valores if v is not None and pd.notna(v) and str(v).strip() != ""]
        if valores_validos:
            total += float(valores_validos[-1])
    return round(total, 1)

def to_float(v):
    if v is None: return None
    try:
        if str(v).strip() == "" or pd.isna(v): return None
        return float(v)
    except Exception:
        return None

def generar_reporte_estados(df_raw: pd.DataFrame, pesos_ra: dict) -> pd.DataFrame:
    if df_raw.empty: return pd.DataFrame()
    reporte = []
    for _, row in df_raw.iterrows():
        estudiante = row['estudiante']
        notas = json.loads(row['notas_json']) if row['notas_json'] else {}
        for ra, peso_total in pesos_ra.items():
            umbral_70 = peso_total * 0.7
            ev = to_float(notas.get(f"{ra} - Ev"))
            r1 = to_float(notas.get(f"{ra} - R1"))
            r2 = to_float(notas.get(f"{ra} - R2"))
            r3 = to_float(notas.get(f"{ra} - R3"))
            
            estado, etapa, nota_vigente = "Sin Evaluar", "-", 0
            if ev is None and r1 is None and r2 is None and r3 is None:
                pass
            else:
                if r3 is not None:
                    estado, etapa, nota_vigente = ("Recuperado", "R3", r3) if r3 >= umbral_70 else ("Reprobado", "R3 (Agotado)", r3)
                elif r2 is not None:
                    estado, etapa, nota_vigente = ("Recuperado", "R2", r2) if r2 >= umbral_70 else ("Pendiente", "R2 (Fallido)", r2)
                elif r1 is not None:
                    estado, etapa, nota_vigente = ("Recuperado", "R1", r1) if r1 >= umbral_70 else ("Pendiente", "R1 (Fallido)", r1)
                elif ev is not None:
                    estado, etapa, nota_vigente = ("Aprobado", "Ev", ev) if ev >= umbral_70 else ("Pendiente", "Ev (Fallido)", ev)
            
            reporte.append({
                "Estudiante": estudiante, "R.A.": ra, "Peso RA": peso_total,
                "Umbral (70%)": round(umbral_70, 1),
                "Nota Vigente": nota_vigente if nota_vigente > 0 else "-",
                "Estatus": estado, "Etapa": etapa
            })
    return pd.DataFrame(reporte)

def obtener_docentes_login() -> pd.DataFrame:
    try:
        conn = sqlite3.connect(DB_NAME)
        df = pd.read_sql_query("SELECT DISTINCT docente, password FROM docentes", conn)
        conn.close()
        return df
    except Exception:
        return pd.DataFrame(columns=["docente", "password"])

def verificar_password(pass_db: str, password: str) -> bool:
    if not pass_db: pass_db = "1234"
    return str(pass_db).strip() == str(password).strip()

def obtener_modulos_docente(docente: str) -> pd.DataFrame:
    try:
        conn = sqlite3.connect(DB_NAME)
        df = pd.read_sql_query("SELECT modulo, seccion FROM docentes WHERE docente = ?", conn, params=(docente,))
        conn.close()
        return df
    except Exception:
        return pd.DataFrame()

def obtener_notas_records(modulo: str, seccion: str) -> dict:
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT estudiante, notas_json FROM calificaciones WHERE modulo=? AND seccion=?", (modulo, seccion))
    rows = cursor.fetchall()
    conn.close()
    return {r[0]: json.loads(r[1]) for r in rows}

def hash_tabla(df: pd.DataFrame) -> str:
    try:
        return hashlib.md5(df.fillna("§NA§").astype(str).to_csv(index=False).encode()).hexdigest()
    except Exception:
        return ""

def guardar_notas_desde_editor(modulo: str, seccion: str, df: pd.DataFrame, pesos_ra: dict) -> None:
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM calificaciones WHERE modulo=? AND seccion=?", (modulo, seccion))
    for _, fila in df.iterrows():
        est = str(fila.get("Estudiante", "")).strip()
        if not est or est.lower() == "none": continue
        notas = {}
        for ra in pesos_ra.keys():
            for intento in INTENTOS:
                k = f"{ra} - {intento}"
                v = fila.get(k)
                try: 
                    notas[k] = None if (v is None or pd.isna(v)) else float(v)
                except Exception: 
                    notas[k] = None
        cursor.execute('INSERT INTO calificaciones (modulo, seccion, estudiante, notas_json) VALUES (?, ?, ?, ?)', 
                       (modulo, seccion, est, json.dumps(notas)))
    conn.commit()
    conn.close()

def importar_excel_matriz(archivo_notas, pesos_ra):
    df_import = pd.read_excel(archivo_notas, header=None, skiprows=3)
    registros = []
    errores = []
    for _, row in df_import.iterrows():
        estudiante = row[1]
        if pd.isna(estudiante) or str(estudiante).strip() == "": continue
        notas_dict = {}
        col_idx = 2
        for ra in pesos_ra.keys():
            for intento in INTENTOS:
                val = row[col_idx]
                try:
                    notas_dict[f"{ra} - {intento}"] = None if pd.isna(val) else int(round(float(val)))
                except Exception:
                    notas_dict[f"{ra} - {intento}"] = None
                col_idx += 1
        registros.append({"estudiante": str(estudiante), "notas": notas_dict})
    return registros, errores

def guardar_notas_registros(modulo: str, seccion: str, registros: list, pesos_ra: dict) -> None:
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM calificaciones WHERE modulo=? AND seccion=?", (modulo, seccion))
    for reg in registros:
        cursor.execute('INSERT INTO calificaciones (modulo, seccion, estudiante, notas_json) VALUES (?, ?, ?, ?)', 
                       (modulo, seccion, reg["estudiante"], json.dumps(reg["notas"])))
    conn.commit()
    conn.close()

def obtener_password_docente(docente: str) -> str:
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT password FROM docentes WHERE docente = ? LIMIT 1", (docente,))
    row = cursor.fetchone()
    conn.close()
    return row[0] if (row and row[0]) else "1234"

def actualizar_password_docente(docente: str, password: str) -> None:
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("UPDATE docentes SET password = ? WHERE docente = ?", (password, docente))
    conn.commit()
    conn.close()

def generar_excel_matriz(df_notas, pesos_ra, seccion_nombre):
    wb = Workbook()
    ws = wb.active
    ws.title = str(seccion_nombre)[:31]

    fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_green = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fill_red = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')
    align_vert = Alignment(horizontal='center', vertical='center', textRotation=90)
    font_bold = Font(bold=True)
    font_bold_large = Font(bold=True, size=16)

    ws.cell(row=1, column=1, value="").fill = fill_gray
    ws.cell(row=1, column=2, value="").fill = fill_yellow
    ws.cell(row=2, column=1, value="").fill = fill_gray
    ws.cell(row=2, column=2, value="").fill = fill_yellow
    
    ws.cell(row=3, column=1, value="#").font = font_bold
    ws.cell(row=3, column=1).fill = fill_gray
    ws.cell(row=3, column=2, value="Nombres").font = font_bold
    ws.cell(row=3, column=2).fill = fill_green

    col_idx = 3
    suma_pesos = 0
    ra_col_ranges = [] 
    
    for ra_name, peso in pesos_ra.items():
        ws.cell(row=1, column=col_idx, value=f"{ra_name} ({peso})").font = font_bold
        ws.cell(row=1, column=col_idx).alignment = align_center
        ws.cell(row=1, column=col_idx).fill = fill_gray
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+3)
        
        ws.cell(row=2, column=col_idx, value=peso).font = font_bold_large
        ws.cell(row=2, column=col_idx).alignment = align_center
        ws.cell(row=2, column=col_idx).fill = fill_gray
        ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx+3)
        
        for i, intento in enumerate(INTENTOS):
            c = ws.cell(row=3, column=col_idx+i, value=intento)
            c.font = Font(bold=True, size=9)
            c.alignment = align_center
            c.fill = fill_gray
            
        for r in range(1, 4):
            for c in range(col_idx, col_idx+4):
                ws.cell(row=r, column=c).border = border_thin
                
        ra_col_ranges.append((col_idx, col_idx+3, peso))
        col_idx += 4
        suma_pesos += peso
        
    col_total = col_idx
    ws.cell(row=1, column=col_total, value="Total").font = font_bold
    ws.cell(row=1, column=col_total).alignment = align_center
    ws.cell(row=1, column=col_total).fill = fill_gray
    ws.merge_cells(start_row=1, start_column=col_total, end_row=1, end_column=col_total)
    
    ws.cell(row=2, column=col_total, value=suma_pesos).font = font_bold_large
    ws.cell(row=2, column=col_total).alignment = align_center
    ws.cell(row=2, column=col_total).fill = fill_gray
    ws.merge_cells(start_row=2, start_column=col_total, end_row=3, end_column=col_total)
    
    col_sit = col_idx + 1
    ws.cell(row=1, column=col_sit, value="Situacion Final").font = Font(bold=True, size=9)
    ws.cell(row=1, column=col_sit).alignment = align_center
    ws.cell(row=1, column=col_sit).fill = fill_gray
    ws.merge_cells(start_row=1, start_column=col_sit, end_row=1, end_column=col_sit+1)
    
    ws.cell(row=2, column=col_sit, value="Aprobado").font = font_bold
    ws.cell(row=2, column=col_sit).alignment = align_vert
    ws.cell(row=2, column=col_sit).fill = fill_green
    ws.merge_cells(start_row=2, start_column=col_sit, end_row=3, end_column=col_sit)
    
    ws.cell(row=2, column=col_sit+1, value="Reprobado").font = font_bold
    ws.cell(row=2, column=col_sit+1).alignment = align_vert
    ws.cell(row=2, column=col_sit+1).fill = fill_red
    ws.merge_cells(start_row=2, start_column=col_sit+1, end_row=3, end_column=col_sit+1)
    
    for r in range(1, 4):
        for c in range(1, col_sit+2):
            if not ws.cell(row=r, column=c).border:
                ws.cell(row=r, column=c).border = border_thin

    row_idx = 4
    num_students = len(df_notas) if not df_notas.empty else 15
    
    for i in range(1, num_students + 1):
        ws.cell(row=row_idx, column=1, value=i).border = border_thin
        ws.cell(row=row_idx, column=1).alignment = align_center
        
        estudiante = ""
        if not df_notas.empty and i-1 < len(df_notas): estudiante = df_notas.iloc[i-1].get('Estudiante', '')
        ws.cell(row=row_idx, column=2, value=estudiante).border = border_thin
        
        c_idx = 3
        if not df_notas.empty and i-1 < len(df_notas):
            row_data = df_notas.iloc[i-1]
            for ra in pesos_ra.keys():
                for intento in INTENTOS:
                    val = row_data.get(f"{ra} - {intento}", "")
                    if pd.isna(val) or val == "": val = ""
                    else: val = int(round(float(val)))
                    c = ws.cell(row=row_idx, column=c_idx, value=val)
                    c.border = border_thin
                    c.alignment = align_center
                    c.number_format = '0'
                    c_idx += 1
        else:
            for _ in range(len(pesos_ra)*4):
                c = ws.cell(row=row_idx, column=c_idx, value="")
                c.border = border_thin
                c.alignment = align_center
                c_idx += 1
                
        formula_parts = []
        for start_c, end_c, _ in ra_col_ranges:
            c_ev = get_column_letter(start_c) + str(row_idx)
            c_r1 = get_column_letter(start_c + 1) + str(row_idx)
            c_r2 = get_column_letter(start_c + 2) + str(row_idx)
            c_r3 = get_column_letter(start_c + 3) + str(row_idx)
            part = f"IF(ISNUMBER({c_r3}),{c_r3},IF(ISNUMBER({c_r2}),{c_r2},IF(ISNUMBER({c_r1}),{c_r1},IF(ISNUMBER({c_ev}),{c_ev},0))))"
            formula_parts.append(part)

        total_formula = "=SUM(" + ",".join(formula_parts) + ")" if formula_parts else "=0"
        c_tot = ws.cell(row=row_idx, column=col_total, value=total_formula)
        c_tot.border = border_thin
        c_tot.alignment = align_center
        c_tot.font = font_bold
        c_tot.number_format = '0'
        
        tot_cell_ref = get_column_letter(col_total) + str(row_idx)
        c_apr = ws.cell(row=row_idx, column=col_sit, value=f'=IF({tot_cell_ref}>=70, "X", "")')
        c_apr.border = border_thin
        c_apr.alignment = align_center
        
        c_rep = ws.cell(row=row_idx, column=col_sit+1, value=f'=IF(AND({tot_cell_ref}<70, {tot_cell_ref}>0), "X", "")')
        c_rep.border = border_thin
        c_rep.alignment = align_center
        
        row_idx += 1

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 35
    for c in range(3, col_total):
        ws.column_dimensions[get_column_letter(c)].width = 5.5
    ws.column_dimensions[get_column_letter(col_total)].width = 8
    ws.column_dimensions[get_column_letter(col_sit)].width = 6
    ws.column_dimensions[get_column_letter(col_sit+1)].width = 6
    ws.row_dimensions[2].height = 45
    ws.row_dimensions[3].height = 20

    fill_rojo_alerta = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    font_roja_alerta = Font(color='9C0006', bold=True)
    
    for start_c, end_c, peso in ra_col_ranges:
        umbral = peso * 0.7
        rule = CellIsRule(operator='lessThan', formula=[str(umbral)], stopIfTrue=True, fill=fill_rojo_alerta, font=font_roja_alerta)
        rango_str = f"{get_column_letter(start_c)}4:{get_column_letter(end_c)}{row_idx-1}"
        ws.conditional_formatting.add(rango_str, rule)
        
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ===========================================================================
# ESTILOS Y COLORES
# ===========================================================================
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; background-color: #F8FAFC; color: #1E293B; }
    .portal-header { background: linear-gradient(135deg, #0F172A 0%, #1E3A8A 100%); color: white; padding: 1.8rem; border-radius: 18px; margin-bottom: 1.25rem; box-shadow: 0 18px 35px rgba(15, 23, 42, 0.18); }
    .portal-title { font-size: 2rem; font-weight: 800; margin-bottom: 0.35rem; }
    .portal-sub { opacity: 0.86; font-size: 1rem; }
    .section-title { color: #1D4ED8; font-weight: 700; font-size: 1.12rem; border-bottom: 2px solid #DBEAFE; padding-bottom: 8px; margin-top: 1.2rem; margin-bottom: 0.9rem; }
    .ra-legend-wrap { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 10px; }
    .ra-chip { display: inline-flex; align-items: center; gap: 6px; padding: 6px 14px; border-radius: 20px; font-weight: 700; font-size: 0.84rem; border: 2px solid; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }
    .ra-chip-peso { font-weight: 500; opacity: 0.75; margin-left: 2px; }
    .grid-caption { color: #64748B; font-size: 0.85rem; margin-bottom: 14px; }
    div.btn-marker-pendiente + div[data-testid="stButton"] button { background-color: #2563EB !important; border-color: #2563EB !important; color: #FFFFFF !important; transition: background-color 0.25s ease; }
    div.btn-marker-pendiente + div[data-testid="stButton"] button:hover { background-color: #1D4ED8 !important; }
    div.btn-marker-guardado + div[data-testid="stButton"] button { background-color: #16A34A !important; border-color: #16A34A !important; color: #FFFFFF !important; transition: background-color 0.25s ease; }
    div.btn-marker-guardado + div[data-testid="stButton"] button:hover { background-color: #15803D !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

PALETA_RA = [
    {"emoji": "🔵", "bg": "#DBEAFE", "border": "#3B82F6", "text": "#1D4ED8"},
    {"emoji": "🟢", "bg": "#D1FAE5", "border": "#10B981", "text": "#065F46"},
    {"emoji": "🟡", "bg": "#FEF9C3", "border": "#EAB308", "text": "#854D0E"},
    {"emoji": "🟣", "bg": "#F3E8FF", "border": "#A855F7", "text": "#6B21A8"},
    {"emoji": "🟠", "bg": "#FFEDD5", "border": "#F97316", "text": "#9A3412"},
    {"emoji": "🔴", "bg": "#FEE2E2", "border": "#EF4444", "text": "#991B1B"},
    {"emoji": "🟤", "bg": "#F0E4D7", "border": "#92400E", "text": "#78350F"},
    {"emoji": "⚫", "bg": "#E2E8F0", "border": "#475569", "text": "#1E293B"},
]

asegurar_tablas()

# ---------------------------------------------------------------------------
# Autenticación integrada con main.py
# ---------------------------------------------------------------------------
def render_login_local() -> None:
    st.markdown(
        """
        <div class="portal-header">
            <div class="portal-title">Portal del Docente ETP</div>
            <div class="portal-sub">Autenticación requerida</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    docentes_df = obtener_docentes_login()
    if docentes_df.empty:
        st.error("⚠️ No hay docentes registrados. Contacta a Coordinación.")
        st.stop()
    with st.form("login_portal_docente"):
        st.write("Selecciona tu nombre e ingresa tu contraseña.")
        docente_sel = st.selectbox("Docente", [""] + sorted(docentes_df["docente"].unique().tolist()))
        password = st.text_input("Contraseña", type="password")
        if st.form_submit_button("🔑 Iniciar sesión", type="primary"):
            if not docente_sel or not password:
                st.error("⚠️ Completa ambos campos.")
            else:
                pass_db = docentes_df.loc[docentes_df["docente"] == docente_sel, "password"].iloc[0]
                if verificar_password(pass_db, password):
                    st.session_state.docente_autenticado = True
                    st.session_state.nombre_docente = docente_sel
                    st.rerun()
                else:
                    st.error("❌ Contraseña incorrecta.")

if not st.session_state.get("docente_autenticado", False):
    render_login_local()
    st.stop()

docente_actual = st.session_state.get("nombre_docente", "")

if not docente_actual:
    st.error("No se pudo identificar al docente autenticado.")
    st.stop()

# ---------------------------------------------------------------------------
# Encabezado
# ---------------------------------------------------------------------------
st.markdown(
    f"""
    <div class="portal-header">
        <div class="portal-title">Portal del Docente ETP</div>
        <div class="portal-sub">Bienvenido/a, <b>{docente_actual}</b>. Gestiona tus calificaciones.</div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# Módulos del docente
# ---------------------------------------------------------------------------
df_asignaciones = obtener_modulos_docente(docente_actual)

if df_asignaciones.empty:
    st.error("No tienes módulos asignados. Consulta con Coordinación.")
    st.stop()

df_asignaciones["Opcion"] = df_asignaciones["modulo"] + " | " + df_asignaciones["seccion"]

opcion_seleccionada = st.selectbox("📚 Selecciona el módulo a gestionar:", df_asignaciones["Opcion"].tolist())

modulo_sel, seccion_sel = opcion_seleccionada.split(" | ")

# ---------------------------------------------------------------------------
# Tabs
# ---------------------------------------------------------------------------
tab_config, tab_notas, tab_import, tab_recuperacion, tab_seguridad = st.tabs(
    ["⚙️ Configurar R.A.", "📝 Matriz de calificaciones", "📥 Importar / Exportar", "🚨 Recuperación", "🔐 Seguridad"]
)

# ---------------------------------------------------------------------------
# Tab 1: Configurar R.A.
# ---------------------------------------------------------------------------
with tab_config:
    st.markdown('<div class="section-title">Ponderación del módulo</div>', unsafe_allow_html=True)
    cant_actual, pesos_actuales = obtener_config_ra(modulo_sel, seccion_sel)

    with st.form(f"form_config_ra_{modulo_sel}_{seccion_sel}"):
        cantidad_ra = st.number_input(
            "Cantidad de Resultados de Aprendizaje (R.A.)",
            min_value=1, max_value=15,
            value=cant_actual if cant_actual > 0 else 5,
        )

        st.markdown("##### Asignar valor a cada R.A.")
        cols = st.columns(5)
        nuevos_pesos = {}
        suma_total = 0

        for i in range(1, int(cantidad_ra) + 1):
            ra_name = f"RA. {i}"
            valor_bd = pesos_actuales.get(ra_name, 0) if pesos_actuales else 0
            valor_defecto = valor_bd if valor_bd > 0 else max(1, 100 // int(cantidad_ra))

            with cols[(i - 1) % 5]:
                val = st.number_input(
                    ra_name, min_value=1, max_value=100, value=valor_defecto,
                    key=f"ra_{modulo_sel}_{seccion_sel}_{i}",
                )
                nuevos_pesos[ra_name] = int(val)
                suma_total += int(val)

        st.progress(min(suma_total, 100) / 100)
        st.info(f"**Suma actual de los R.A.: {suma_total} / 100 puntos**")

        if st.form_submit_button("💾 Guardar configuración", type="primary"):
            if suma_total != 100:
                st.error("❌ Los valores deben sumar exactamente 100 puntos.")
            else:
                guardar_config_ra(modulo_sel, seccion_sel, cantidad_ra, nuevos_pesos)
                st.success("✅ Configuración guardada correctamente.")
                st.rerun()

# ---------------------------------------------------------------------------
# Tab 2: Matriz de calificaciones
# ---------------------------------------------------------------------------
with tab_notas:
    st.markdown('<div class="section-title">Matriz oficial de calificaciones</div>', unsafe_allow_html=True)
    cant_actual, pesos_actuales = obtener_config_ra(modulo_sel, seccion_sel)

    if cant_actual == 0:
        st.warning("⚠️ Configura los R.A. en la pestaña anterior.")
    else:
        ra_keys = list(pesos_actuales.keys())

        with st.container(border=True):
            chips_html = '<div class="ra-legend-wrap">'
            for idx, ra in enumerate(ra_keys):
                pal = PALETA_RA[idx % len(PALETA_RA)]
                chips_html += (
                    f'<span class="ra-chip" style="'
                    f"background:{pal['bg']};border-color:{pal['border']};color:{pal['text']};\">"
                    f"{pal['emoji']} {ra} <span class='ra-chip-peso'>({pesos_actuales[ra]} pts)</span></span>"
                )
            chips_html += "</div>"
            st.markdown(chips_html, unsafe_allow_html=True)

            st.markdown(
                """
                <div class="grid-caption">
                <b>Ev</b> = Evaluación inicial · <b>R1/R2/R3</b> = Recuperaciones.
                Edita directamente las celdas y luego guarda.
                </div>
                """,
                unsafe_allow_html=True,
            )

            db_records = obtener_notas_records(modulo_sel, seccion_sel)
            nombres_existentes = list(db_records.keys())
            num_filas = max(len(nombres_existentes), 12)
            filas = []

            for i in range(num_filas):
                nombre = nombres_existentes[i] if i < len(nombres_existentes) else ""
                notas_est = db_records.get(nombre, {}) if nombre else {}
                fila = {"Estudiante": nombre}

                for idx, ra in enumerate(ra_keys):
                    pal = PALETA_RA[idx % len(PALETA_RA)]
                    fila[f"{ra}__sep"] = f"{pal['emoji']} {ra}"
                    for intento in INTENTOS:
                        fila[f"{ra} - {intento}"] = to_float(notas_est.get(f"{ra} - {intento}"))
                filas.append(fila)

            columnas = ["Estudiante"]
            for ra in ra_keys:
                columnas.append(f"{ra}__sep")
                columnas.extend(f"{ra} - {intento}" for intento in INTENTOS)

            df_editor_base = pd.DataFrame(filas, columns=columnas)
            column_config = {"Estudiante": st.column_config.TextColumn("👤 Nombres", width="medium")}

            for idx, ra in enumerate(ra_keys):
                pal = PALETA_RA[idx % len(PALETA_RA)]
                peso_ra = pesos_actuales[ra]
                column_config[f"{ra}__sep"] = st.column_config.TextColumn(
                    f"{pal['emoji']} {ra}", width="small", disabled=True,
                    help=f"Bloque de {ra} (vale {peso_ra} pts).",
                )
                for intento in INTENTOS:
                    column_config[f"{ra} - {intento}"] = st.column_config.NumberColumn(
                        intento, min_value=0, max_value=peso_ra, step=1, format="%d",
                        help=f"Nota de {intento} para {ra} (máximo {peso_ra} pts).",
                    )

            editor_key = f"editor_notas_{modulo_sel}_{seccion_sel}"
            edited_df = st.data_editor(
                df_editor_base, column_config=column_config, num_rows="dynamic",
                use_container_width=True, hide_index=True, height=460, key=editor_key,
            )

            df_resumen = edited_df.copy()
            df_resumen["Nota Final"] = df_resumen.apply(lambda row: calcular_nota_final_fila(row, pesos_actuales), axis=1)
            df_resumen = df_resumen[df_resumen["Estudiante"].astype(str).str.strip() != ""]

            if not df_resumen.empty:
                st.markdown("###### 🧮 Nota Final (vista previa en vivo)")
                st.dataframe(df_resumen[["Estudiante", "Nota Final"]], use_container_width=True, hide_index=True, height=180)

            hash_actual = hash_tabla(edited_df)
            hash_key = f"hash_guardado_notas_{modulo_sel}_{seccion_sel}"
            flag_key = f"guardado_ok_notas_{modulo_sel}_{seccion_sel}"

            if hash_key not in st.session_state: st.session_state[hash_key] = None
            if flag_key not in st.session_state: st.session_state[flag_key] = False
            if st.session_state[hash_key] != hash_actual: st.session_state[flag_key] = False

            guardado_ok = st.session_state[flag_key]
            marker_class = "btn-marker-guardado" if guardado_ok else "btn-marker-pendiente"
            texto_boton = "✅ ¡Guardado correctamente!" if guardado_ok else "💾 Guardar calificaciones"

            st.markdown(f'<div class="{marker_class}"></div>', unsafe_allow_html=True)

            if st.button(texto_boton, use_container_width=True, key=f"btn_guardar_notas_{modulo_sel}_{seccion_sel}"):
                guardar_notas_desde_editor(modulo_sel, seccion_sel, edited_df, pesos_actuales)
                st.session_state[hash_key] = hash_actual
                st.session_state[flag_key] = True
                st.success("✅ ¡Calificaciones guardadas exitosamente!")
                st.rerun()

# ---------------------------------------------------------------------------
# Tab 3: Importar / Exportar
# ---------------------------------------------------------------------------
with tab_import:
    st.markdown('<div class="section-title">Importar / Exportar matriz</div>', unsafe_allow_html=True)
    cant_actual, pesos_actuales = obtener_config_ra(modulo_sel, seccion_sel)

    if cant_actual == 0:
        st.warning("Configura los R.A. primero.")
    else:
        col_export, col_import = st.columns(2)

        with col_export:
            st.markdown("#### 📤 Descargar matriz actual")
            df_notas_db = obtener_notas_df(modulo_sel, seccion_sel)
            df_excel = construir_df_notas(df_notas_db)
            buffer = generar_excel_matriz(df_excel, pesos_actuales, seccion_sel)

            st.download_button(
                "📥 Descargar Excel MINERD",
                data=buffer,
                file_name=f"Matriz_{modulo_sel}_{seccion_sel}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

        with col_import:
            st.markdown("#### 📥 Importar matriz llenada")
            archivo_notas = st.file_uploader("Sube el archivo Excel", type=["xlsx"])

            if archivo_notas:
                records_key = f"import_records_{modulo_sel}_{seccion_sel}"
                errors_key = f"import_errors_{modulo_sel}_{seccion_sel}"
                file_key = f"import_file_{modulo_sel}_{seccion_sel}"

                if st.session_state.get(file_key) != archivo_notas.name or records_key not in st.session_state:
                    try:
                        registros, errores = importar_excel_matriz(archivo_notas, pesos_actuales)
                        st.session_state[records_key] = registros
                        st.session_state[errors_key] = errores
                        st.session_state[file_key] = archivo_notas.name
                    except Exception as e:
                        st.error(f"❌ Error leyendo el Excel: {e}")
                        registros, errores = [], []
                else:
                    registros = st.session_state.get(records_key, [])
                    errores = st.session_state.get(errors_key, [])

                if errores:
                    with st.expander(f"⚠️ Advertencias ({len(errores)})"):
                        for error in errores: st.markdown(f"- {error}")

                if registros:
                    st.markdown("##### Vista previa")
                    st.dataframe(pd.DataFrame(registros).head(20), use_container_width=True, hide_index=True)
                    if st.button("🚀 Confirmar importación", type="primary", use_container_width=True):
                        guardar_notas_registros(modulo_sel, seccion_sel, registros, pesos_actuales)
                        st.success("✅ Matriz importada correctamente.")
                        st.rerun()
                else:
                    st.info("No se detectaron registros válidos en el Excel.")

# ---------------------------------------------------------------------------
# Tab 4: Recuperación
# ---------------------------------------------------------------------------
with tab_recuperacion:
    st.markdown('<div class="section-title">Seguimiento de recuperación</div>', unsafe_allow_html=True)
    cant_actual, pesos_actuales = obtener_config_ra(modulo_sel, seccion_sel)

    if cant_actual == 0:
        st.warning("Configura los R.A. primero.")
    else:
        df_notas_db = obtener_notas_df(modulo_sel, seccion_sel)
        df_reporte = generar_reporte_estados(df_notas_db, pesos_actuales)

        if df_reporte.empty:
            st.info("No hay calificaciones registradas para evaluar estatus.")
        else:
            filtro = st.selectbox("Filtrar por estatus", ["Todos", "Recuperado", "Pendiente", "Aprobado", "Reprobado", "Sin Evaluar"])
            df_rep = df_reporte.copy()

            if filtro != "Todos":
                df_rep = df_rep[df_rep["Estatus"].str.contains(filtro, case=False, na=False)]

            c1, c2, c3, c4 = st.columns(4)
            with c1: st.metric("Registros", len(df_rep))
            with c2: st.metric("Recuperados", int(df_rep["Estatus"].str.contains("Recuperado", na=False).sum()))
            with c3: st.metric("Pendientes", int(df_rep["Estatus"].str.contains("Pendiente", na=False).sum()))
            with c4: st.metric("Reprobados", int(df_rep["Estatus"].str.contains("Reprobado", na=False).sum()))

            st.markdown("#### Distribución por estatus")
            st.bar_chart(df_rep["Estatus"].value_counts())
            st.dataframe(df_rep, use_container_width=True, hide_index=True)

            buffer_rep = BytesIO()
            df_rep.to_excel(buffer_rep, index=False, engine="openpyxl")
            buffer_rep.seek(0)

            st.download_button(
                "📥 Descargar reporte de estatus (.xlsx)", data=buffer_rep,
                file_name=f"Estatus_Recuperacion_{modulo_sel}_{seccion_sel}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ---------------------------------------------------------------------------
# Tab 5: Seguridad
# ---------------------------------------------------------------------------
with tab_seguridad:
    st.markdown('<div class="section-title">🔐 Cambiar mi contraseña</div>', unsafe_allow_html=True)

    with st.form(f"form_seguridad_{docente_actual}"):
        pass_actual = st.text_input("Contraseña actual", type="password")
        pass_nueva = st.text_input("Nueva contraseña", type="password")
        pass_confirma = st.text_input("Confirmar nueva contraseña", type="password")

        if st.form_submit_button("Actualizar contraseña", type="primary"):
            pass_db = obtener_password_docente(docente_actual)

            if not verificar_password(pass_db, pass_actual):
                st.error("❌ La contraseña actual es incorrecta.")
            elif pass_nueva != pass_confirma:
                st.error("❌ Las nuevas contraseñas no coinciden.")
            elif len(pass_nueva) < 6:
                st.error("❌ La nueva contraseña debe tener al menos 6 caracteres.")
            else:
                actualizar_password_docente(docente_actual, pass_nueva)
                st.success("✅ Contraseña actualizada correctamente.")