"""
directorio_docentes.py — Directorio de Docentes ETP
Adaptado para ecosistema main.py (Standalone)
KPIs, editor, importación de plantilla oficial y exportación Excel/CSV.
"""
import datetime
import re
import unicodedata
import sqlite3
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ═══════════════════════════════════════════════════════════════════════════
# Base de Datos (Integración con gestion_etp.db)
# ═══════════════════════════════════════════════════════════════════════════
DB_NAME = "gestion_etp.db"

def asegurar_esquema():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS docentes (
        docente TEXT, modulo TEXT, seccion TEXT, password TEXT DEFAULT '1234', usuario TEXT
    )''')
    cols = [info[1] for info in cur.execute("PRAGMA table_info(docentes)").fetchall()]
    if 'area_tecnica' not in cols: cur.execute("ALTER TABLE docentes ADD COLUMN area_tecnica TEXT")
    if 'horas' not in cols: cur.execute("ALTER TABLE docentes ADD COLUMN horas TEXT")
    conn.commit()
    conn.close()

def obtener_directorio():
    conn = sqlite3.connect(DB_NAME)
    df = pd.read_sql_query("SELECT docente, modulo, seccion, area_tecnica, horas FROM docentes", conn)
    conn.close()
    return df.to_dict('records')

def listar_docentes():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT docente FROM docentes WHERE docente IS NOT NULL AND docente != ''")
    rows = [r[0] for r in cur.fetchall()]
    conn.close()
    return rows

def reemplazar_directorio(registros):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    
    # Guardar credenciales existentes para no perderlas al actualizar
    cur.execute("SELECT docente, password, usuario FROM docentes WHERE docente IS NOT NULL AND docente != ''")
    creds = {}
    for r in cur.fetchall():
        if r[0] not in creds:
            creds[r[0]] = {"password": r[1], "usuario": r[2]}
            
    cur.execute("DELETE FROM docentes")
    
    count = 0
    for reg in registros:
        docente = reg["docente"]
        modulo = reg["modulo"]
        seccion = reg["seccion"]
        area = reg["area_tecnica"]
        horas = reg.get("horas", "")
        
        pwd = creds.get(docente, {}).get("password", "1234")
        usu = creds.get(docente, {}).get("usuario", str(docente).split()[0].lower() if docente else "docente")
        
        cur.execute("INSERT INTO docentes (docente, modulo, seccion, password, usuario, area_tecnica, horas) VALUES (?,?,?,?,?,?,?)",
                    (docente, modulo, seccion, pwd, usu, area, str(horas) if horas else ""))
        count += 1
        
    conn.commit()
    conn.close()
    return count

def eliminar_todos_usuarios():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("DELETE FROM docentes")
    count = cur.rowcount
    conn.commit()
    conn.close()
    return count

def vaciar_modulos():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT docente, password, usuario FROM docentes")
    docentes = cur.fetchall()
    cur.execute("DELETE FROM docentes")
    for d in docentes:
        cur.execute("INSERT INTO docentes (docente, modulo, seccion, password, usuario, area_tecnica, horas) VALUES (?, '', '', ?, ?, '', '')", (d[0], d[1], d[2]))
    count = len(docentes)
    conn.commit()
    conn.close()
    return count

def restablecer_password(docente, nueva_clave):
    if len(str(nueva_clave)) < 4:
        raise ValueError("La contraseña debe tener al menos 4 caracteres.")
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("UPDATE docentes SET password=? WHERE docente=?", (nueva_clave, docente))
    conn.commit()
    conn.close()


COL_OFICIALES = ["Docente Asignado", "Área Técnica", "Módulo Formativo", "Sección", "Horas del Módulo"]

# ═══════════════════════════════════════════════════════════════════════════
# Estilos
# ═══════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; background-color: #F8FAFC; color: #1E293B; }
.dir-hero { background: linear-gradient(135deg, #0F172A 0%, #1E3A8A 60%, #2563EB 100%); color: #fff;
    padding: 1.6rem 1.8rem; border-radius: 16px; margin-bottom: 1.1rem; box-shadow: 0 18px 35px rgba(15,23,42,.18); }
.dir-title { font-size: 1.9rem; font-weight: 800; letter-spacing: -0.02em; }
.dir-sub { opacity: .85; font-size: .95rem; margin-top: .3rem; }
.section-title { color: #1D4ED8; font-weight: 700; font-size: 1.1rem; border-bottom: 2px solid #DBEAFE;
    padding-bottom: 8px; margin: 1.1rem 0 .9rem 0; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════
# Guardia + datos
# ═══════════════════════════════════════════════════════════════════════════
if not st.session_state.get("coordinador_autenticado", False):
    st.error("🔒 Esta página es exclusiva de Coordinación.")
    st.stop()

asegurar_esquema()


def _df_desde_auth() -> pd.DataFrame:
    registros = obtener_directorio()
    df = pd.DataFrame(registros)
    if df.empty:
        return pd.DataFrame(columns=COL_OFICIALES)
    df = df.rename(columns={
        "docente": "Docente Asignado", "area_tecnica": "Área Técnica",
        "modulo": "Módulo Formativo", "seccion": "Sección", "horas": "Horas del Módulo",
    })
    df = df[df["Módulo Formativo"].astype(str).str.strip() != ""]
    return df.reindex(columns=COL_OFICIALES, fill_value="")


def _registros_desde_df(df: pd.DataFrame) -> list[dict]:
    registros = []
    for _, row in df.iterrows():
        docente = str(row.get("Docente Asignado") or "").strip()
        modulo = str(row.get("Módulo Formativo") or "").strip()
        seccion = str(row.get("Sección") or "").strip()
        if not docente or not modulo or not seccion:
            continue
        try:
            horas = int(float(row.get("Horas del Módulo")))
        except Exception:
            horas = None
        registros.append({
            "docente": docente, "area_tecnica": str(row.get("Área Técnica") or "").strip(),
            "modulo": modulo, "seccion": seccion, "horas": horas,
        })
    return registros


def _normalizar_encabezado(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())


MAPA_ENCABEZADOS = {
    "docenteasignado": "Docente Asignado", "areatecnica": "Área Técnica",
    "moduloformativo": "Módulo Formativo", "seccion": "Sección",
    "horasdelmodulo": "Horas del Módulo",
}


def calcular_kpis(df: pd.DataFrame) -> dict:
    horas = pd.to_numeric(df["Horas del Módulo"], errors="coerce").fillna(0)
    return {
        "docentes": int(df["Docente Asignado"].nunique()),
        "areas": int(df["Área Técnica"].nunique()),
        "modulos": int(len(df)),
        "modulos_unicos": int(df["Módulo Formativo"].nunique()),
        "secciones": int(df["Sección"].nunique()),
        "horas": int(horas.sum()),
    }


def _tarjeta_kpi(icono, valor, etiqueta, detalle, tono) -> str:
    return f"""
    <div style="background:#fff;border:1px solid #E2E8F0;border-top:4px solid {tono};border-radius:12px;
        padding:12px 16px;box-shadow:0 4px 12px rgba(15,23,42,.06);height:100%;">
      <div style="display:flex;justify-content:space-between;align-items:center;">
        <span style="font-size:.72rem;font-weight:800;color:#64748B;text-transform:uppercase;letter-spacing:.05em;">{etiqueta}</span>
        <span style="font-size:1.15rem;">{icono}</span>
      </div>
      <div style="font-size:1.9rem;font-weight:800;color:#0F172A;margin:.15rem 0;">{valor}</div>
      <div style="font-size:.78rem;color:#94A3B8;">{detalle}</div>
    </div>"""


def generar_excel_directorio(df: pd.DataFrame) -> BytesIO:
    """Excel formateado: hoja Directorio + hoja Resumen (carga por docente/área)."""
    k = calcular_kpis(df)
    wb = Workbook()
    ws = wb.active
    ws.title = "Directorio"
    fill_hdr = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    font_hdr = Font(bold=True, color="FFFFFF", size=10)
    fill_zebra = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    borde = Border(left=Side(style="thin", color="D7DEE8"), right=Side(style="thin", color="D7DEE8"),
                   top=Side(style="thin", color="D7DEE8"), bottom=Side(style="thin", color="D7DEE8"))
    centro = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:E1")
    ws["A1"] = "Directorio de Docentes ETP — Carga Horaria Oficial"
    ws["A1"].font = Font(bold=True, size=13, color="1F3A5F")
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(size=9, color="64748B", italic=True)

    for col_idx, nombre in enumerate(COL_OFICIALES, start=1):
        c = ws.cell(row=3, column=col_idx, value=nombre)
        c.fill, c.font, c.alignment, c.border = fill_hdr, font_hdr, centro, borde
    for r_idx, (_, row) in enumerate(df.iterrows(), start=4):
        for col_idx, nombre in enumerate(COL_OFICIALES, start=1):
            valor = row[nombre]
            c = ws.cell(row=r_idx, column=col_idx,
                        value=(int(valor) if nombre == "Horas del Módulo" and pd.notna(valor) else valor))
            c.border = borde
            if nombre in ("Sección", "Horas del Módulo"):
                c.alignment = centro
            if r_idx % 2 == 0:
                c.fill = fill_zebra
    anchos = {"A": 32, "B": 40, "C": 48, "D": 12, "E": 12}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A3:E{max(3, len(df) + 3)}"
    ws.freeze_panes = "A4"

    # ── Hoja Resumen ──
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Resumen Ejecutivo del Directorio"
    ws2["A1"].font = Font(bold=True, size=13, color="1F3A5F")
    resumen = [
        ("Docentes únicos", k["docentes"]), ("Áreas técnicas únicas", k["areas"]),
        ("Módulos asignados", k["modulos"]), ("Módulos sin repetición", k["modulos_unicos"]),
        ("Secciones únicas", k["secciones"]), ("Horas semanales totales", k["horas"]),
    ]
    for i, (label, val) in enumerate(resumen, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=val).font = Font(bold=True, color="1D4ED8")
    fila = len(resumen) + 5
    ws2.cell(row=fila, column=1, value="Carga por docente").font = Font(bold=True, size=11, color="1F3A5F")
    fila += 1
    for col_idx, h in enumerate(["Docente", "Módulos", "Horas"], start=1):
        c = ws2.cell(row=fila, column=col_idx, value=h)
        c.fill, c.font = fill_hdr, font_hdr
    grp = df.groupby("Docente Asignado").agg(
        modulos=("Módulo Formativo", "size"),
        horas=("Horas del Módulo", lambda s: int(pd.to_numeric(s, errors="coerce").fillna(0).sum())),
    ).reset_index().sort_values("horas", ascending=False)
    for r_idx, (_, row) in enumerate(grp.iterrows(), start=fila + 1):
        ws2.cell(row=r_idx, column=1, value=row["Docente Asignado"])
        ws2.cell(row=r_idx, column=2, value=int(row["modulos"])).alignment = centro
        ws2.cell(row=r_idx, column=3, value=int(row["horas"])).alignment = centro
    fila2 = fila + len(grp) + 3
    ws2.cell(row=fila2, column=1, value="Distribución por área técnica").font = Font(bold=True, size=11, color="1F3A5F")
    fila2 += 1
    for col_idx, h in enumerate(["Área Técnica", "Docentes", "Módulos", "Horas"], start=1):
        c = ws2.cell(row=fila2, column=col_idx, value=h)
        c.fill, c.font = fill_hdr, font_hdr
    grp_a = df.groupby("Área Técnica").agg(
        docentes=("Docente Asignado", "nunique"), modulos=("Módulo Formativo", "size"),
        horas=("Horas del Módulo", lambda s: int(pd.to_numeric(s, errors="coerce").fillna(0).sum())),
    ).reset_index()
    for r_idx, (_, row) in enumerate(grp_a.iterrows(), start=fila2 + 1):
        ws2.cell(row=r_idx, column=1, value=row["Área Técnica"])
        ws2.cell(row=r_idx, column=2, value=int(row["docentes"])).alignment = centro
        ws2.cell(row=r_idx, column=3, value=int(row["modulos"])).alignment = centro
        ws2.cell(row=r_idx, column=4, value=int(row["horas"])).alignment = centro
    ws2.column_dimensions["A"].width = 45
    for col in "BCD":
        ws2.column_dimensions[col].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ═══════════════════════════════════════════════════════════════════════════
# Encabezado + KPIs
# ═══════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="dir-hero">
    <div class="dir-title">🧑‍🏫 Directorio de Docentes ETP</div>
    <div class="dir-sub">Carga horaria, módulos y accesos · sincronizado con usuarios</div>
</div>
""", unsafe_allow_html=True)

df_actual = _df_desde_auth()
kpis = calcular_kpis(df_actual)

c1, c2, c3, c4, c5 = st.columns(5)
with c1:
    st.markdown(_tarjeta_kpi("🧑🏫", kpis["docentes"], "Docentes", "sin repetición", "#2563EB"), unsafe_allow_html=True)
with c2:
    st.markdown(_tarjeta_kpi("🏷️", kpis["areas"], "Áreas técnicas", "sin repetición", "#7C3AED"), unsafe_allow_html=True)
with c3:
    st.markdown(_tarjeta_kpi("📚", kpis["modulos"], "Módulos", f"{kpis['modulos_unicos']} sin repetición", "#0EA5E9"), unsafe_allow_html=True)
with c4:
    st.markdown(_tarjeta_kpi("🏫", kpis["secciones"], "Secciones", "sin repetición", "#F59E0B"), unsafe_allow_html=True)
with c5:
    st.markdown(_tarjeta_kpi("⏱️", kpis["horas"], "Horas semanales", "suma de carga", "#10B981"), unsafe_allow_html=True)

tab1, tab2, tab3, tab4, tab5 = st.tabs(
    ["🔍 Consultas", "✏️ Editor", "📥 Importar Excel", "📤 Exportar", "🔐 Accesos y Claves"]
)

# ═══════════════════════════════════════════════════════════════════════════
# Tab 1: Consultas
# ═══════════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown('<div class="section-title">Filtros y búsqueda</div>', unsafe_allow_html=True)
    if df_actual.empty:
        st.info("Aún no hay docentes registrados. Usa el Editor o importa la plantilla Excel.")
    else:
        busqueda = st.text_input("🔎 Búsqueda global", placeholder="Docente, módulo, área o sección…")
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            f_doc = st.multiselect("Docente(s)", sorted(df_actual["Docente Asignado"].unique()))
        with col_f2:
            f_area = st.multiselect("Área(s) técnica(s)", sorted(df_actual["Área Técnica"].unique()))
        with col_f3:
            f_sec = st.multiselect("Sección(es)", sorted(df_actual["Sección"].unique()))
        df_f = df_actual.copy()
        if busqueda.strip():
            t = busqueda.strip().lower()
            df_f = df_f[df_f.apply(lambda r: t in " ".join(str(v).lower() for v in r), axis=1)]
        if f_doc:
            df_f = df_f[df_f["Docente Asignado"].isin(f_doc)]
        if f_area:
            df_f = df_f[df_f["Área Técnica"].isin(f_area)]
        if f_sec:
            df_f = df_f[df_f["Sección"].isin(f_sec)]
        kf = calcular_kpis(df_f)
        st.caption(
            f"**{len(df_f)}** asignaciones · **{kf['docentes']}** docentes · "
            f"**{kf['secciones']}** secciones · **{kf['horas']}** horas"
        )
        st.dataframe(df_f, width="stretch", hide_index=True)
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            st.download_button(
                "📥 Exportar vista filtrada (.xlsx)", data=generar_excel_directorio(df_f),
                file_name="Directorio_filtrado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
        with col_d2:
            st.download_button(
                "📄 Exportar vista filtrada (.csv)", data=df_f.to_csv(index=False).encode("utf-8-sig"),
                file_name="Directorio_filtrado.csv", mime="text/csv", width="stretch",
            )
        st.divider()
        st.markdown('<div class="section-title">Ficha de carga por docente</div>', unsafe_allow_html=True)
        doc_sel = st.selectbox("Selecciona un docente", [""] + sorted(df_actual["Docente Asignado"].unique()))
        if doc_sel:
            df_doc = df_actual[df_actual["Docente Asignado"] == doc_sel]
            horas_doc = int(pd.to_numeric(df_doc["Horas del Módulo"], errors="coerce").fillna(0).sum())
            m1, m2, m3 = st.columns(3)
            m1.metric("Módulos", len(df_doc))
            m2.metric("Secciones", df_doc["Sección"].nunique())
            m3.metric("Horas semanales", horas_doc)
            st.dataframe(df_doc.drop(columns=["Docente Asignado"]), width="stretch", hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════
# Tab 2: Editor
# ═══════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="section-title">Modificación de registros</div>', unsafe_allow_html=True)
    st.info("💡 Edita, agrega o elimina filas directamente. **Guardar** sincroniza usuarios y módulos sin tocar contraseñas.")
    areas_existentes = sorted({a for a in df_actual["Área Técnica"].dropna() if str(a).strip()})
    col_config = {
        "Docente Asignado": st.column_config.TextColumn("Docente Asignado", required=True),
        "Área Técnica": st.column_config.SelectboxColumn(
            "Área Técnica",
            options=sorted(set(["Informática", "Electromecánica", "Contabilidad", "Redes", "Logística",
                                "Servicios Turísticos", "Comercio y Marketing", "Otra"] + areas_existentes)),
        ),
        "Módulo Formativo": st.column_config.TextColumn("Módulo Formativo", required=True),
        "Sección": st.column_config.TextColumn("Sección (Ej. 4to A)", required=True),
        "Horas del Módulo": st.column_config.NumberColumn("Horas", min_value=1, max_value=60, step=1),
    }
    edited_df = st.data_editor(df_actual, column_config=col_config, num_rows="dynamic",
                               width="stretch", hide_index=True)
    if st.button("💾 Guardar cambios", type="primary"):
        registros = _registros_desde_df(edited_df)
        if not registros:
            st.warning("⚠️ No hay registros válidos (falta docente, módulo o sección).")
        else:
            previos = set(listar_docentes())
            n = reemplazar_directorio(registros)
            nuevos = set(listar_docentes()) - previos
            msg = f"✅ Directorio actualizado: {n} asignaciones."
            if nuevos:
                msg += f" Usuarios nuevos (clave 1234): {', '.join(sorted(nuevos))}."
            st.toast(msg, icon="✅")
            st.rerun()

# ═══════════════════════════════════════════════════════════════════════════
# Tab 3: Importar Excel
# ═══════════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="section-title">Carga masiva de la plantilla oficial</div>', unsafe_allow_html=True)
    st.info("📂 Formato esperado: *Docente Asignado · Área Técnica · Módulo Formativo · Sección · Horas del Módulo*. "
            "Reconoce encabezados aunque vengan con encoding dañado.")
    archivo_subido = st.file_uploader("Sube el archivo Excel", type=["xlsx", "xls"])
    if archivo_subido is not None:
        try:
            df_imp = pd.read_excel(archivo_subido)
            df_imp.columns = [MAPA_ENCABEZADOS.get(_normalizar_encabezado(c), str(c)) for c in df_imp.columns]
            df_imp = df_imp[[c for c in COL_OFICIALES if c in df_imp.columns]]
            k_imp = calcular_kpis(df_imp)
            st.caption(
                f"Detectados: **{k_imp['docentes']}** docentes · **{k_imp['areas']}** áreas · "
                f"**{k_imp['modulos']}** módulos · **{k_imp['secciones']}** secciones · **{k_imp['horas']}** horas"
            )
            st.dataframe(df_imp, width="stretch", hide_index=True)

            st.markdown("#### 🧹 Opciones de reemplazo")
            reemplazo_total = st.checkbox(
                "Eliminar los datos actuales antes de importar (reemplazo completo)",
                help="Si no lo marcas, la importación solo actualiza/crea los docentes de la plantilla y conserva el resto.",
            )
            alcance_borrado = None
            if reemplazo_total:
                alcance_borrado = st.radio(
                    "Alcance del borrado",
                    ["Conservar usuarios y claves (borrar solo módulos)",
                     "Reinicio total (borrar usuarios, módulos y claves)"],
                    help="El reinicio total recrea los usuarios de la plantilla con clave 1234.",
                )
                st.warning(
                    "⚠️ Esta acción no se puede deshacer. Los docentes que NO estén en la plantilla perderán su carga horaria"
                    + (" y sus usuarios." if alcance_borrado.startswith("Reinicio") else ".")
                )

            if st.button("🚀 Importar y sincronizar", type="primary"):
                registros = _registros_desde_df(df_imp)
                if not registros:
                    st.error("❌ La plantilla no contiene registros válidos.")
                else:
                    if reemplazo_total:
                        if alcance_borrado and alcance_borrado.startswith("Reinicio"):
                            n_borrado = eliminar_todos_usuarios()
                        else:
                            n_borrado = vaciar_modulos()
                    previos = set(listar_docentes())
                    n = reemplazar_directorio(registros)
                    nuevos = set(listar_docentes()) - previos
                    msg = f"✅ Importación completada: {n} asignaciones."
                    if nuevos:
                        msg += f"\n🆕 Usuarios creados (clave 1234): {', '.join(sorted(nuevos))}."
                    st.toast(msg, icon="✅")
                    st.rerun()
        except Exception as e:
            st.error(f"⚠️ Error leyendo el Excel: {e}")

    st.divider()
    with st.expander("🗑️ Zona de peligro: vaciar directorio sin importar"):
        st.warning("⚠️ Elimina la carga horaria actual **sin** cargar una plantilla nueva. Úsalo solo para empezar de cero.")
        confirmar_vaciado = st.checkbox("Confirmo que deseo vaciar el directorio", key="chk_vaciar")
        alcance_vaciado = st.radio(
            "Alcance",
            ["Borrar solo módulos (conservar usuarios)", "Borrar usuarios y módulos (reinicio total)"],
            key="radio_vaciar",
        )
        if st.button("🗑️ Vaciar directorio ahora", type="primary",
                     disabled=not confirmar_vaciado, key="btn_vaciar"):
            if alcance_vaciado.startswith("Borrar solo"):
                n = vaciar_modulos()
                detalle = f"vaciado de módulos ({n} asignaciones)"
            else:
                n = eliminar_todos_usuarios()
                detalle = f"reinicio total ({n} usuarios eliminados)"
            st.toast(f"✅ Directorio vaciado: {detalle}.", icon="🗑️")
            st.rerun()

# ═══════════════════════════════════════════════════════════════════════════
# Tab 4: Exportar
# ═══════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown('<div class="section-title">Exportación institucional</div>', unsafe_allow_html=True)
    if df_actual.empty:
        st.info("No hay datos para exportar.")
    else:
        col_e1, col_e2 = st.columns(2)
        with col_e1:
            st.markdown("#### 📊 Excel formateado")
            st.caption("Hoja *Directorio* con filtros y formatos + hoja *Resumen* con conteos sin repetición, "
                       "carga por docente y distribución por área.")
            st.download_button(
                "📥 Descargar Excel completo (.xlsx)", data=generar_excel_directorio(df_actual),
                file_name="Directorio_Docentes_ETP.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch", type="primary",
            )
        with col_e2:
            st.markdown("#### 📄 CSV (UTF-8 con BOM)")
            st.caption("Compatible con Excel en español: abre sin perder acentos ni tildes.")
            st.download_button(
                "📥 Descargar CSV (.csv)", data=df_actual.to_csv(index=False).encode("utf-8-sig"),
                file_name="Directorio_Docentes_ETP.csv", mime="text/csv", width="stretch",
            )

# ═══════════════════════════════════════════════════════════════════════════
# Tab 5: Accesos y Claves
# ═══════════════════════════════════════════════════════════════════════════
with tab5:
    st.markdown('<div class="section-title">🔐 Restablecer contraseñas de docentes</div>', unsafe_allow_html=True)
    st.write("Selecciona el docente y asígnale una clave provisional.")
    docentes_lista = listar_docentes()
    if not docentes_lista:
        st.warning("No hay usuarios en la base de datos.")
    else:
        col_r1, col_r2 = st.columns([2, 1])
        with col_r1:
            doc_reset = st.selectbox("Seleccionar docente", sorted(docentes_lista))
        with col_r2:
            nueva_pass = st.text_input("Nueva contraseña temporal", value="1234")
        if st.button("🔄 Restablecer contraseña", type="primary"):
            try:
                restablecer_password(doc_reset, nueva_pass)
                st.toast(f"✅ Contraseña de {doc_reset} restablecida.", icon="🔐")
            except ValueError as e:
                st.error(f"⚠️ {e}")